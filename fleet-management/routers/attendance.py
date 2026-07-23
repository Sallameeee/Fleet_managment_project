"""Attendance reports — MANAGER ONLY (dashboard), school orgs only.

Exports student attendance to Excel (openpyxl) or PDF (reportlab), with a
manager-chosen set/order of columns. Two shapes, both via the same endpoint:
  * ALL students over a route + date range (omit student_id), or
  * a MONTHLY report for ONE student (pass student_id + the month's date range).

Data-source: the `attendance` table (written by the supervisor app), joined to
the student's profile/route. Org-scoped to the caller; refuses non-school orgs.
"""

import io
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status

from auth import require_permission
from capacity_logic import LOCAL_TZ
from database import supabase

router = APIRouter(prefix="/attendance", tags=["attendance"])

# Column key -> header label. The manager picks which of these to include + order.
COLUMN_LABELS = {
    "student_name": "Student",
    "class_name": "Class",
    "grade": "Grade",
    "route_name": "Route",
    "date": "Date",
    "session": "Session",          # morning (pickup) / afternoon (drop-off)
    "boarded": "Boarded",
    "drop_off_stop": "Drop-off stop",  # where the student got off (afternoon)
    "time": "Time",
    "parent_phone": "Parent phone",
    "student_phone": "Student phone",
}
DEFAULT_COLUMNS = ["student_name", "class_name", "route_name", "date", "session", "boarded", "drop_off_stop", "time"]


def _local_hhmm(recorded_at) -> Optional[str]:
    """`recorded_at` is stored UTC (timestamptz). Render it as HH:MM in the org's
    LOCAL time (Africa/Cairo, DST-aware) — slicing the raw ISO string showed UTC,
    which read an hour or two early on the sheet."""
    if not recorded_at:
        return None
    try:
        dt = datetime.fromisoformat(str(recorded_at).replace("Z", "+00:00"))
        if dt.tzinfo is None:  # naive value from the DB → it is UTC
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(LOCAL_TZ).strftime("%H:%M")
    except Exception:
        return str(recorded_at)[11:16] or None


def _dedupe_latest(att: list) -> list:
    """One row per (student, date, session) — the LATEST recording wins.

    Attendance is upserted per (trip, student), so a student riding TWO trips in
    the same session that day (e.g. the supervisor restarted the trip) yields two
    rows and the sheet showed the child twice. The raw per-trip rows are kept in
    the DB (audit); only the REPORT collapses them."""
    best: dict = {}
    for a in att:
        key = (a.get("student_id"), a.get("trip_date"), a.get("session"))
        cur = best.get(key)
        if cur is None or str(a.get("recorded_at") or "") > str(cur.get("recorded_at") or ""):
            best[key] = a
    return list(best.values())


def _require_school_org(org_id: str) -> None:
    module = "university"
    try:
        r = supabase.table("organizations").select("module").eq("id", org_id).limit(1).execute()
        if r.data and r.data[0].get("module"):
            module = r.data[0]["module"]
    except Exception:
        pass
    if module != "school":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Attendance reports are only available for school organizations.",
        )


def _fetch_rows(
    org_id: str,
    route_id: Optional[str],
    student_id: Optional[str],
    date_from: Optional[date],
    date_to: Optional[date],
    attendance_status: str = "both",
) -> list:
    """Enriched attendance rows (one per student per DATE+SESSION) for the filters.

    `attendance_status`: 'present' (boarded only) | 'absent' (not boarded) | 'both'.
    """
    q = supabase.table("attendance").select("student_id, trip_id, trip_date, boarded, session, drop_off_stop, recorded_at").eq("org_id", org_id)
    if student_id:
        q = q.eq("student_id", student_id)
    if date_from:
        q = q.gte("trip_date", date_from.isoformat())
    if date_to:
        q = q.lte("trip_date", date_to.isoformat())
    att = q.order("trip_date", desc=False).execute().data

    # Collapse repeat trips in the same session, THEN apply the present/absent
    # filter so the filter sees each student's final status for that session.
    att = _dedupe_latest(att)
    if attendance_status == "present":
        att = [a for a in att if a.get("boarded")]
    elif attendance_status == "absent":
        att = [a for a in att if not a.get("boarded")]

    sids = list({a["student_id"] for a in att})
    pax = {}
    if sids:
        pax = {
            p["id"]: p
            for p in supabase.table("passengers")
            .select("id, name, route_id, grade, class_name, parent_phone, student_phone")
            .in_("id", sids)
            .execute()
            .data
        }

    # The route on the SHEET is the bus the student ACTUALLY rode that record —
    # i.e. the route of the TRIP the attendance was recorded on. On a one-day bus
    # change this is the NEW bus, not the student's home route. Fall back to the
    # student's profile route only when a record has no resolvable trip.
    trip_ids = list({a.get("trip_id") for a in att if a.get("trip_id")})
    trip_route = {}
    if trip_ids:
        trip_route = {
            t["id"]: t.get("route_id")
            for t in supabase.table("trips").select("id, route_id").in_("id", trip_ids).execute().data
        }

    def _actual_route_id(a):
        return trip_route.get(a.get("trip_id")) or (pax.get(a["student_id"]) or {}).get("route_id")

    if route_id:  # filter to one route — by the bus actually ridden that day
        att = [a for a in att if _actual_route_id(a) == route_id]
        sids = list({a["student_id"] for a in att})

    route_ids = list({_actual_route_id(a) for a in att if _actual_route_id(a)})
    routes = {}
    if route_ids:
        routes = {r["id"]: r["name"] for r in supabase.table("routes").select("id, name").in_("id", route_ids).execute().data}

    rows = []
    for a in att:
        p = pax.get(a["student_id"]) or {}
        rows.append(
            {
                "student_name": p.get("name"),
                "class_name": p.get("class_name"),
                "grade": p.get("grade"),
                "route_name": routes.get(_actual_route_id(a)),
                "date": a.get("trip_date"),
                "session": (a.get("session") or "").capitalize() or None,
                "boarded": "Yes" if a.get("boarded") else "No",
                "drop_off_stop": a.get("drop_off_stop"),
                "time": _local_hhmm(a.get("recorded_at")),
                "parent_phone": p.get("parent_phone"),
                "student_phone": p.get("student_phone"),
            }
        )
    rows.sort(key=lambda x: (x["date"] or "", (x["student_name"] or "").lower(), x["session"] or ""))
    return rows


def _render_xlsx(headers: list, rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, 1):
        longest = max([len(str(h))] + [len(str(r[i - 1])) for r in rows]) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(i)].width = min(40, max(12, longest + 2))
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_pdf(headers: list, rows: list, title: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Attendance Report")
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    if not rows:
        story.append(Paragraph("No attendance records for this selection.", styles["Italic"]))
    else:
        table = Table([headers] + rows, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f7")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(table)
    doc.build(story)
    return buf.getvalue()


@router.get("/export")
def export_attendance(
    current_user: dict = Depends(require_permission("manage_passengers")),
    fmt: str = Query("xlsx", alias="format", description="xlsx | pdf"),
    columns: Optional[str] = Query(None, description="Comma-separated column keys, in order."),
    route_id: Optional[str] = Query(None),
    student_id: Optional[str] = Query(None, description="Set for a single-student (e.g. monthly) report."),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    attendance_status: str = Query(
        "both",
        alias="status",
        description="Which records to include: present (boarded) | absent (not boarded) | both.",
    ),
):
    org_id = current_user["org_id"]
    _require_school_org(org_id)

    attendance_status = (attendance_status or "both").lower()
    if attendance_status not in ("present", "absent", "both"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="status must be one of: present, absent, both.",
        )

    cols = [c for c in (columns.split(",") if columns else DEFAULT_COLUMNS) if c in COLUMN_LABELS]
    if not cols:
        cols = DEFAULT_COLUMNS
    headers = [COLUMN_LABELS[c] for c in cols]

    data = _fetch_rows(org_id, route_id, student_id, date_from, date_to, attendance_status)
    table = [[("" if r.get(c) is None else str(r.get(c))) for c in cols] for r in data]

    title = "Attendance report"
    if student_id and data:
        title = f"Attendance — {data[0]['student_name'] or 'Student'}"

    if fmt == "pdf":
        content = _render_pdf(headers, table, title)
        media = "application/pdf"
        ext = "pdf"
    else:
        content = _render_xlsx(headers, table)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    fname = f"attendance_{date_from or 'all'}_{date_to or 'all'}.{ext}"
    return Response(
        content=content,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/columns")
def attendance_columns(current_user: dict = Depends(require_permission("manage_passengers"))):
    """The available export columns (key + label) for the manager's column picker."""
    _require_school_org(current_user["org_id"])
    return {"columns": [{"key": k, "label": v} for k, v in COLUMN_LABELS.items()], "default": DEFAULT_COLUMNS}
